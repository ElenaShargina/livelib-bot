import logging
import datetime

import datetime as datetime
import os.path
import string

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from .parser import BookDataFormatter, ParserForXLSX
from .config import Config


class Export:
    """
    Абстрактный класс для экспорта книг читателя в файл.
    """
    folder: str = ''  # папка, где будут сохраняться экспортируемые файлы
    encoding: str = ''  # кодировка, в которой будут сохраняться экспортируемые файлы

    def create_file(self, books: list, login: str) -> str or None:
        """
        Сохраняет файл с заданными книгами для читателя с заданным логином.
        Логин будет использован в названии файла.
        :type books: list
        :type login: str
        :return: путь к сохраненному файлу
        :rtype str:
        """
    pass

class XLSXExport(Export):
    """
    Класс для экспорта списка книг читателя в файл .xlsx
    """

    def __init__(self, config: Config):
        self.folder = config.export.xlsx.folder
        self.encoding = config.encoding

    def _create_filename(self, reader_name: str) -> str:
        """
        Служебный метод возвращает имя файла экспорта для читателя с добавленной датой экспорта.
        Считаем, что имя читателя берется из его логина в БД, поэтому безопасно для использования в имени файла.
        """
        if reader_name == '' or None:
            logging.exception('При экспорте не был указан логин читателя!')
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d--%H-%M')
        filename = timestamp + '-' + reader_name + '.xlsx'
        return os.path.join(self.folder, filename).replace('\\', '/')

    def create_file(self, books: list, login: str, parser_xlsx: ParserForXLSX) -> str or None:
        """
        Сохраняет файл с заданными книгами в формате xlsx для читателя с заданным логином.
        Логин будет использован в названии файла.
        :type books: list
        :type login: str
        :return: путь к сохраненному файлу
        :rtype str:
        """
        # узнаем словарь с описанием колонок в экспортном файле
        properties = BookDataFormatter.all_properties_xlsx()
        # формируем путь и название экспортного файла
        filename = self._create_filename(login)
        # создаем документ Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Прочитанные книги'

        # Задаем ширину столбцов. Она берется из BookDataFormatter
        index = string.ascii_uppercase
        i = 0
        for value in properties.values():
            if value.get('column_width', None) and i < 26:
                ws.column_dimensions[index[i]].width = value['column_width']
                i += 1

        # Задаем названия столбцов
        title_row = [i['name'] for i in properties.values()]
        ws.append(title_row)
        #  Задаем стиль для первой строки
        for i in range(1, 1 + len(properties)):
            ws.cell(ws.max_row, i).style = 'Headline 2'

        # Вводим данные про все книги
        for book in books:
            prepared_book = parser_xlsx.prepare_book_for_xlsx(book)
            # print(prepared_book)
            ws.append(list(prepared_book.values()))
            # свойства-ссылки вводим как ссылки
            for link_property in ('author_id', 'book_id', 'work_id', 'picture_url', 'review_id'):
                ws.cell(ws.max_row, properties[link_property]['order']).hyperlink = prepared_book[link_property]
                ws.cell(ws.max_row, properties[link_property]['order']).style = 'Hyperlink'
            # если есть рецензия, то делаем вертикальное выравнивание
            if prepared_book['review_text']:
                # ws.row_dimensions[ws.max_row].height = 100
                for i in range(1, 1 + len(properties)):
                    ws.cell(ws.max_row, i).alignment = Alignment(vertical='center')
                ws.cell(ws.max_row, properties['review_text']['order']).alignment = Alignment(wrapText=True,
                                                                                              vertical='center')
        # Сохраняем файл
        try:
            wb.save(filename)
            logging.info(
                f'Successfully saved! Export file with {len(books)} books in xlsx for reader {login} at {filename}')
            return filename
        except Exception as exc:
            logging.exception(f'Не удалось сохранить файл c экспортом по адресу {filename}')
            return None