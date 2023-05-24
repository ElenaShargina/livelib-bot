import datetime
import json
import os, sys

# скрипт для правильной отработки тестов в github.actions
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils import get_correct_filename, CustomUnitTest, remove_file, create_logger_for_tests
import unittest
import logging
import random
import openpyxl
from livelib import *


class TestXLSXExport(CustomUnitTest):
    # папка с данными для проведения тестов
    # внутри её есть подпапки для конкретных методов
    test_folder: str = 'data/sample/test_export/xlsx'
    # файл с конфигом для проведения тестов.
    # тесты имеют собственную папку кеша веб-страниц
    config_file: str = '.env.export'

    @classmethod
    def setUpClass(cls) -> None:
        create_logger_for_tests(__name__+'.log')
        cls.config = Config(get_correct_filename(cls.config_file, ''))
        cls.config.web_connection.cache_folder = get_correct_filename('',cls.config.web_connection.cache_folder)
        cls.web_connection = WebWithCache(cls.config, random_sleep=False)
        cls.config.db.sqlite_db = get_correct_filename(cls.config.db.sqlite_db, "")
        cls.db_connection = SQLite3Connection(cls.config, create_if_not_exist=False)
        cls.parser = ParserFromHTML
        cls.config.export.xlsx.folder = get_correct_filename('',cls.config.export.xlsx.folder)
        cls.export = XLSXExport(cls.config)

    def test__create_filename(self):
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d--%H-%M')
        values = [{'reader_name': 'some_reader', 'folder': 'xlsx', 'output': f"xlsx/{timestamp}-some_reader.xlsx"},
                  {'reader_name': 'Foo', 'folder': 'xlsx/somefolder',
                   'output': f"xlsx/somefolder/{timestamp}-Foo.xlsx"},
                  {'reader_name': 'some_reader', 'folder': '', 'output': f"{timestamp}-some_reader.xlsx"}]
        for i in values:
            special_config = Config(get_correct_filename(self.config_file, ''))
            special_config.export.xlsx.folder = i['folder']
            export = XLSXExport(special_config)
            self.assertEqual(i['output'], export._create_filename(i['reader_name']).replace('\\','/'))

    def test_create_file(self):
        self.test_folder = get_correct_filename('',os.path.join(self.test_folder, 'create_file'))
        for reader_name in ['Eugenia_Novik', 'Humming_Bird']:
            # 1. Создаем нового читателя
            # Формируем особый конфиг, чтобы кеш страниц брался из подготовленных данных.
            # Если страница реального читателя поменяется, то сравнение в тесте будет все равно идти с сохраненной старой версией.
            special_config = self.config
            special_config.web_connection.cache_folder = get_correct_filename("",'data/sample/test_export/xlsx/create_file/cache')
            my_reader = Reader(reader_name, WebWithCache(special_config), self.db_connection, self.export)
            my_reader.register()
            my_reader.get_read_books_from_web()

            # 2. Формируем экспортный файл
            output_filename = my_reader.create_export_xlsx_file()
            # print(output_filename)

            # 3. Сравниваем с образцом
            # загружаем полученный ранее вариант
            my_output = openpyxl.reader.excel.load_workbook(output_filename, read_only=True)
            wb1 = my_output.active
            values = [i for i in wb1.values]
            my_output.close()

            # загружаем образец
            correct_output_filename = os.path.join(self.test_folder,reader_name+'-correct_output.xlsx')
            # print(correct_output_filename)
            correct_output = openpyxl.reader.excel.load_workbook(correct_output_filename, read_only = True)
            # correct_output = openpyxl.reader.excel.load_workbook(output_filename, read_only=True)
            wb = correct_output.active
            correct_values = [i for i in wb.values]
            correct_output.close()

            with self.subTest(f'Testing create_file with {reader_name}'):
                self.assertEqual(correct_values, values)

            # 5. Удаляем данные
            # удаляем книги читателя из БД
            my_reader.delete_read_books()
            # удаляем файл экспорта
            remove_file(output_filename)
